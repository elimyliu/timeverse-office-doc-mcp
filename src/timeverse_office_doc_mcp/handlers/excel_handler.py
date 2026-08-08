"""Excel 文档处理器 - 24 个工具。

对应方案 5.2 Excel 工具集。使用 openpyxl + pandas 实现。
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries

from ..common.error_handler import ToolError
from ..common.file_lock import file_lock_mgr
from ..common.path_guard import path_guard
from ..common.session import session_manager
from ..common.template_mgr import template_manager
from ..common.validator import InputValidator

logger = logging.getLogger("timeverse_office_doc_mcp.excel")


# ==================== 辅助函数 ====================


def _get_workbook(filename: str, session_id: str | None = None) -> Workbook:
    """获取 Workbook 对象：Session 模式从内存取，否则从磁盘打开。"""
    if session_id:
        session = session_manager.get_session(session_id)
        if session.format != "excel":
            raise ToolError(f"Session {session_id} 不是 Excel 文档")
        return session.document
    validated = path_guard.validate_path(filename, "read")
    return load_workbook(validated)


def _save_workbook(wb: Workbook, filename: str, session_id: str | None = None) -> None:
    """保存工作簿：Session 模式仅标记修改，否则写入磁盘。"""
    if session_id:
        session_manager.mark_modified(session_id)
    else:
        validated = path_guard.validate_path(filename, "write")
        file_lock_mgr.acquire(validated)
        try:
            wb.save(validated)
        finally:
            file_lock_mgr.release(validated)


def _get_sheet(wb: Workbook, sheet_name: str) -> Any:
    """获取工作表，不存在则报错。"""
    if sheet_name not in wb.sheetnames:
        raise ToolError(f"工作表不存在: {sheet_name}（可用: {wb.sheetnames}）")
    return wb[sheet_name]


# ==================== 5.2.1 工作簿管理（7 个） ====================


def excel_create_workbook(
    filename: str,
    sheet_name: str = "Sheet",
    template: str | None = None,
    variables: dict[str, Any] | None = None,
    session_id: str | None = None,
) -> dict[str, Any]:
    """创建新工作簿（支持模板）。"""
    InputValidator.validate_filename(filename)
    validated = path_guard.validate_path(filename, "write")

    if template:
        fmt, tpl_path = template_manager.resolve_template_path(template)
        if fmt != "excel":
            raise ToolError(f"模板 '{template}' 是 {fmt} 格式，不是 excel")
        wb = load_workbook(tpl_path)
        replaced = _fill_template_variables(wb, variables) if variables else 0
    else:
        wb = Workbook()
        if wb.sheetnames:
            wb.active.title = sheet_name
        replaced = 0

    if session_id:
        session_manager.open_session(validated, "excel", wb)
    else:
        file_lock_mgr.acquire(validated)
        try:
            wb.save(validated)
        finally:
            file_lock_mgr.release(validated)

    return {
        "filename": validated,
        "template": template,
        "variables_replaced": replaced,
        "sheets": wb.sheetnames,
    }


def _fill_template_variables(wb: Workbook, variables: dict[str, Any]) -> int:
    """填充工作簿中的 {{variable}} 占位符。"""
    import re

    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "{{" in cell.value:

                    def replacer(m: re.Match[str]) -> str:
                        return str(variables.get(m.group(1), ""))

                    new_val = re.sub(r"\{\{([^}]+)\}\}", replacer, cell.value)
                    if new_val != cell.value:
                        cell.value = new_val
                        count += 1
    return count


def excel_get_info(filename: str, session_id: str | None = None) -> dict[str, Any]:
    """获取工作簿信息。"""
    wb = _get_workbook(filename, session_id)
    sheets_info = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets_info.append(
            {
                "name": name,
                "rows": ws.max_row,
                "cols": ws.max_column,
            }
        )
    return {
        "filename": filename,
        "sheet_count": len(wb.sheetnames),
        "sheets": sheets_info,
    }


def excel_list_sheets(filename: str, session_id: str | None = None) -> dict[str, Any]:
    """列出工作表。"""
    wb = _get_workbook(filename, session_id)
    return {"filename": filename, "sheets": wb.sheetnames}


def excel_add_sheet(
    filename: str, sheet_name: str, session_id: str | None = None
) -> dict[str, Any]:
    """添加工作表。"""
    wb = _get_workbook(filename, session_id)
    if sheet_name in wb.sheetnames:
        raise ToolError(f"工作表已存在: {sheet_name}")
    wb.create_sheet(title=sheet_name)
    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "sheet_name": sheet_name, "sheets": wb.sheetnames}


def excel_delete_sheet(
    filename: str, sheet_name: str, session_id: str | None = None
) -> dict[str, Any]:
    """删除工作表。"""
    wb = _get_workbook(filename, session_id)
    _get_sheet(wb, sheet_name)
    if len(wb.sheetnames) <= 1:
        raise ToolError("不能删除最后一个工作表")
    wb.remove(wb[sheet_name])
    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "deleted": sheet_name, "sheets": wb.sheetnames}


def excel_rename_sheet(
    filename: str, old_name: str, new_name: str, session_id: str | None = None
) -> dict[str, Any]:
    """重命名工作表。"""
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, old_name)
    if new_name in wb.sheetnames:
        raise ToolError(f"工作表名已存在: {new_name}")
    ws.title = new_name
    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "old_name": old_name, "new_name": new_name}


def excel_copy_sheet(
    filename: str, source: str, target: str, session_id: str | None = None
) -> dict[str, Any]:
    """复制工作表。"""
    wb = _get_workbook(filename, session_id)
    src_ws = _get_sheet(wb, source)
    if target in wb.sheetnames:
        raise ToolError(f"目标工作表已存在: {target}")
    new_ws = wb.copy_worksheet(src_ws)
    new_ws.title = target
    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "source": source, "target": target}


# ==================== 5.2.2 数据读写（8 个） ====================


def excel_read_cell(
    filename: str, sheet: str, cell_ref: str, session_id: str | None = None
) -> dict[str, Any]:
    """读取单元格。"""
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)
    cell = ws[cell_ref]
    return {"filename": filename, "sheet": sheet, "cell": cell_ref, "value": cell.value}


def excel_write_cell(
    filename: str,
    sheet: str,
    cell_ref: str,
    value: Any,
    session_id: str | None = None,
) -> dict[str, Any]:
    """写入单元格。"""
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)
    ws[cell_ref] = value
    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "sheet": sheet, "cell": cell_ref, "value": value}


def excel_read_range(
    filename: str, sheet: str, range_str: str, session_id: str | None = None
) -> dict[str, Any]:
    """读取区域数据。"""
    InputValidator.validate_range(range_str)
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)
    cell_range = ws[range_str]
    data = [[cell.value for cell in row] for row in cell_range]
    return {"filename": filename, "sheet": sheet, "range": range_str, "data": data}


def excel_write_range(
    filename: str,
    sheet: str,
    start_cell: str,
    data: list[list[Any]],
    session_id: str | None = None,
) -> dict[str, Any]:
    """批量写入区域。"""
    if not data:
        raise ToolError("数据不能为空")
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)
    min_col, min_row, _, _ = range_boundaries(f"{start_cell}:{start_cell}")
    for r_idx, row_data in enumerate(data):
        for c_idx, value in enumerate(row_data):
            ws.cell(row=min_row + r_idx, column=min_col + c_idx, value=value)
    _save_workbook(wb, filename, session_id)
    return {
        "filename": filename,
        "sheet": sheet,
        "start_cell": start_cell,
        "rows_written": len(data),
        "cols_written": len(data[0]) if data else 0,
    }


def excel_insert_row(
    filename: str,
    sheet: str,
    row_idx: int,
    count: int = 1,
    session_id: str | None = None,
) -> dict[str, Any]:
    """插入行。"""
    InputValidator.validate_positive_int(row_idx, "row_idx")
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)
    ws.insert_rows(row_idx, count)
    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "sheet": sheet, "row_idx": row_idx, "count": count}


def excel_delete_row(
    filename: str,
    sheet: str,
    row_idx: int,
    count: int = 1,
    session_id: str | None = None,
) -> dict[str, Any]:
    """删除行。"""
    InputValidator.validate_positive_int(row_idx, "row_idx")
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)
    ws.delete_rows(row_idx, count)
    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "sheet": sheet, "row_idx": row_idx, "count": count}


def excel_insert_column(
    filename: str,
    sheet: str,
    col_idx: int,
    count: int = 1,
    session_id: str | None = None,
) -> dict[str, Any]:
    """插入列。"""
    InputValidator.validate_positive_int(col_idx, "col_idx")
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)
    ws.insert_cols(col_idx, count)
    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "sheet": sheet, "col_idx": col_idx, "count": count}


def excel_delete_column(
    filename: str,
    sheet: str,
    col_idx: int,
    count: int = 1,
    session_id: str | None = None,
) -> dict[str, Any]:
    """删除列。"""
    InputValidator.validate_positive_int(col_idx, "col_idx")
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)
    ws.delete_cols(col_idx, count)
    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "sheet": sheet, "col_idx": col_idx, "count": count}


# ==================== 5.2.3 格式化与高级（7 个） ====================


def excel_format_cell(
    filename: str,
    sheet: str,
    range_str: str,
    font: str | None = None,
    bold: bool = False,
    italic: bool = False,
    font_size: int | None = None,
    font_color: str | None = None,
    bg_color: str | None = None,
    alignment: str | None = None,
    border_style: str | None = None,
    session_id: str | None = None,
) -> dict[str, Any]:
    """格式化单元格。"""
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)

    font_obj = Font(
        name=font,
        bold=bold,
        italic=italic,
        size=font_size,
        color=font_color,
    )
    fill_obj = (
        PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        if bg_color
        else None
    )
    align_obj = Alignment(horizontal=alignment) if alignment else None

    border_obj = None
    if border_style:
        side = Side(style=border_style)
        border_obj = Border(left=side, right=side, top=side, bottom=side)

    for row in ws[range_str]:
        for cell in row:
            cell.font = font_obj
            if fill_obj:
                cell.fill = fill_obj
            if align_obj:
                cell.alignment = align_obj
            if border_obj:
                cell.border = border_obj

    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "sheet": sheet, "range": range_str}


def excel_apply_formula(
    filename: str,
    sheet: str,
    cell_ref: str,
    formula: str,
    session_id: str | None = None,
) -> dict[str, Any]:
    """应用公式。"""
    if not formula.startswith("="):
        formula = "=" + formula
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)
    ws[cell_ref] = formula
    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "sheet": sheet, "cell": cell_ref, "formula": formula}


def excel_create_chart(
    filename: str,
    sheet: str,
    chart_type: str,
    data_range: str,
    title: str = "",
    session_id: str | None = None,
) -> dict[str, Any]:
    """创建图表。"""
    InputValidator.validate_choice(chart_type, ["bar", "line", "pie"], "chart_type")
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)

    chart_classes = {"bar": BarChart, "line": LineChart, "pie": PieChart}
    chart = chart_classes[chart_type]()

    min_col, min_row, max_col, max_row = range_boundaries(data_range)
    data_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    chart.add_data(ws[data_ref], titles_from_data=True)

    if title:
        chart.title = title

    anchor_cell = f"{get_column_letter(max_col + 2)}{min_row}"
    ws.add_chart(chart, anchor_cell)

    _save_workbook(wb, filename, session_id)
    return {
        "filename": filename,
        "sheet": sheet,
        "chart_type": chart_type,
        "data_range": data_range,
        "title": title,
        "anchor": anchor_cell,
    }


def excel_freeze_panes(
    filename: str, sheet: str, cell_ref: str, session_id: str | None = None
) -> dict[str, Any]:
    """冻结窗格。"""
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)
    ws.freeze_panes = cell_ref
    _save_workbook(wb, filename, session_id)
    return {"filename": filename, "sheet": sheet, "freeze_at": cell_ref}


def excel_sort_data(
    filename: str,
    sheet: str,
    range_str: str,
    key_column: int,
    ascending: bool = True,
    session_id: str | None = None,
) -> dict[str, Any]:
    """排序数据。"""
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)

    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    rows_data = []
    for r in range(min_row, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        rows_data.append(row_vals)

    rows_data.sort(
        key=lambda x: (x[key_column - 1] is None, x[key_column - 1]), reverse=not ascending
    )

    for r_idx, row_vals in enumerate(rows_data):
        for c_idx, val in enumerate(row_vals):
            ws.cell(row=min_row + r_idx, column=min_col + c_idx, value=val)

    _save_workbook(wb, filename, session_id)
    return {
        "filename": filename,
        "sheet": sheet,
        "range": range_str,
        "key_column": key_column,
        "ascending": ascending,
        "rows_sorted": len(rows_data),
    }


def excel_create_pivot_table(
    filename: str,
    source_sheet: str,
    source_range: str,
    target_sheet: str,
    rows: str,
    cols: str,
    values: str,
    agg_func: str = "sum",
    session_id: str | None = None,
) -> dict[str, Any]:
    """创建数据透视表（用 pandas 实现）。"""
    import pandas as pd

    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, source_sheet)

    min_col, min_row, max_col, max_row = range_boundaries(source_range)
    data_rows = []
    headers = []
    for c in range(min_col, max_col + 1):
        headers.append(str(ws.cell(row=min_row, column=c).value))
    for r in range(min_row + 1, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        data_rows.append(row_vals)

    df = pd.DataFrame(data_rows, columns=headers)
    agg_map = {"sum": "sum", "mean": "mean", "count": "count", "max": "max", "min": "min"}
    agg = agg_map.get(agg_func, "sum")
    pivot = df.pivot_table(index=rows, columns=cols, values=values, aggfunc=agg, fill_value=0)

    if target_sheet not in wb.sheetnames:
        wb.create_sheet(title=target_sheet)
    target_ws = wb[target_sheet]

    target_ws.cell(row=1, column=1, value=f"{rows} \\ {cols}")
    for c_idx, col_name in enumerate(pivot.columns, 2):
        target_ws.cell(row=1, column=c_idx, value=str(col_name))
    for r_idx, (idx_name, row_data) in enumerate(pivot.iterrows(), 2):
        target_ws.cell(row=r_idx, column=1, value=str(idx_name))
        for c_idx, val in enumerate(row_data, 2):
            target_ws.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else 0)

    _save_workbook(wb, filename, session_id)
    return {
        "filename": filename,
        "target_sheet": target_sheet,
        "rows": rows,
        "cols": cols,
        "values": values,
        "agg_func": agg_func,
        "pivot_rows": len(pivot),
        "pivot_cols": len(pivot.columns),
    }


def excel_add_conditional_format(
    filename: str,
    sheet: str,
    range_str: str,
    rule_type: str,
    criteria: str | None = None,
    format_color: str | None = None,
    session_id: str | None = None,
) -> dict[str, Any]:
    """添加条件格式。"""
    InputValidator.validate_choice(
        rule_type, ["greater_than", "less_than", "equal", "between", "contains_text"], "rule_type"
    )
    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)

    fill = PatternFill(
        start_color=format_color or "FF0000", end_color=format_color or "FF0000", fill_type="solid"
    )

    if rule_type == "greater_than" and criteria:
        rule = CellIsRule(operator="greaterThan", formula=[criteria], fill=fill)
    elif rule_type == "less_than" and criteria:
        rule = CellIsRule(operator="lessThan", formula=[criteria], fill=fill)
    elif rule_type == "equal" and criteria:
        rule = CellIsRule(operator="equal", formula=[criteria], fill=fill)
    elif rule_type == "between" and criteria:
        parts = criteria.split(",")
        if len(parts) != 2:
            raise ToolError("between 规则需要 'min,max' 格式的 criteria")
        rule = CellIsRule(
            operator="between", formula=[parts[0].strip(), parts[1].strip()], fill=fill
        )
    elif rule_type == "contains_text" and criteria:
        from openpyxl.formatting.rule import FormulaRule

        rule = FormulaRule(formula=[f'NOT(ISERROR(SEARCH("{criteria}",A1)))'], fill=fill)
    else:
        raise ToolError(f"规则类型 '{rule_type}' 需要 criteria 参数")

    ws.conditional_formatting.add(range_str, rule)
    _save_workbook(wb, filename, session_id)
    return {
        "filename": filename,
        "sheet": sheet,
        "range": range_str,
        "rule_type": rule_type,
        "criteria": criteria,
    }


# ==================== 5.2.4 分析工具（2 个） ====================


def excel_analyze_data(
    filename: str,
    sheet: str,
    range_str: str | None = None,
    session_id: str | None = None,
) -> dict[str, Any]:
    """数据统计分析（描述统计、空值检测、类型推断）。"""
    import pandas as pd

    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)

    if range_str:
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
    else:
        min_col, min_row, max_col, max_row = 1, 1, ws.max_column, ws.max_row

    headers = [
        str(ws.cell(row=min_row, column=c).value or f"Col{c}") for c in range(min_col, max_col + 1)
    ]
    data_rows = []
    for r in range(min_row + 1, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        data_rows.append(row_vals)

    df = pd.DataFrame(data_rows, columns=headers)

    stats = {}
    for col in df.columns:
        col_data = df[col]
        null_count = col_data.isna().sum()
        type_inferred = "unknown"
        numeric_data = pd.to_numeric(col_data, errors="coerce")
        if numeric_data.notna().sum() > len(col_data) * 0.5:
            type_inferred = "numeric"
            stats[col] = {
                "type": type_inferred,
                "null_count": int(null_count),
                "count": int(col_data.notna().sum()),
                "mean": float(numeric_data.mean()) if numeric_data.notna().any() else None,
                "min": float(numeric_data.min()) if numeric_data.notna().any() else None,
                "max": float(numeric_data.max()) if numeric_data.notna().any() else None,
                "std": float(numeric_data.std()) if numeric_data.notna().sum() > 1 else None,
            }
        else:
            type_inferred = "text"
            unique_count = col_data.nunique()
            stats[col] = {
                "type": type_inferred,
                "null_count": int(null_count),
                "count": int(col_data.notna().sum()),
                "unique_count": int(unique_count),
                "most_common": str(col_data.mode().iloc[0]) if not col_data.mode().empty else None,
            }

    return {
        "filename": filename,
        "sheet": sheet,
        "range": range_str or "all",
        "total_rows": len(data_rows),
        "total_cols": len(headers),
        "columns": stats,
    }


def excel_find_duplicates(
    filename: str,
    sheet: str,
    columns: list[str] | None = None,
    threshold: int = 1,
    session_id: str | None = None,
) -> dict[str, Any]:
    """查找重复数据。"""
    import pandas as pd

    wb = _get_workbook(filename, session_id)
    ws = _get_sheet(wb, sheet)

    headers = [
        str(ws.cell(row=1, column=c).value or f"Col{c}") for c in range(1, ws.max_column + 1)
    ]
    data_rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        data_rows.append(row_vals)

    df = pd.DataFrame(data_rows, columns=headers)

    target_cols = columns if columns else headers
    dup_mask = df.duplicated(subset=target_cols, keep=False)
    duplicates = df[dup_mask]

    dup_groups = (
        duplicates.groupby(target_cols)
        .size()
        .reset_index(name="count")
        .sort_values("count", ascending=False)
    )

    return {
        "filename": filename,
        "sheet": sheet,
        "columns_checked": target_cols,
        "total_duplicates": int(len(duplicates)),
        "duplicate_groups": int(len(dup_groups)),
        "details": dup_groups.head(20).to_dict("records") if not dup_groups.empty else [],
    }
