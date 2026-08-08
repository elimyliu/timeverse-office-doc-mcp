"""测试 Excel handler 工具。"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from timeverse_office_doc_mcp.handlers import excel_handler


@pytest.fixture
def workspace(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """创建临时工作空间并更新 PathGuard 白名单。"""
    ws = tmp_path / "workspace"
    ws.mkdir()
    from timeverse_office_doc_mcp.common.path_guard import path_guard

    monkeypatch.setattr(path_guard, "allowed_dirs", [tmp_path.resolve()])
    return ws


@pytest.fixture
def xlsx_path(workspace: Path) -> str:
    return str(workspace / "test.xlsx")


class TestExcelCreateWorkbook:
    def test_create_empty(self, xlsx_path: str) -> None:
        result = excel_handler.excel_create_workbook(xlsx_path, sheet_name="数据")
        assert result["filename"] == str(Path(xlsx_path).resolve())
        assert "数据" in result["sheets"]
        assert Path(xlsx_path).exists()

    def test_create_and_read_back(self, xlsx_path: str) -> None:
        excel_handler.excel_create_workbook(xlsx_path)
        wb = load_workbook(xlsx_path)
        assert "Sheet" in wb.sheetnames


class TestExcelSheetOperations:
    def test_add_and_delete_sheet(self, xlsx_path: str) -> None:
        excel_handler.excel_create_workbook(xlsx_path)
        excel_handler.excel_add_sheet(xlsx_path, "分析")
        info = excel_handler.excel_get_info(xlsx_path)
        sheet_names = [s["name"] for s in info["sheets"]]
        assert "分析" in sheet_names

        excel_handler.excel_delete_sheet(xlsx_path, "分析")
        info = excel_handler.excel_get_info(xlsx_path)
        sheet_names = [s["name"] for s in info["sheets"]]
        assert "分析" not in sheet_names

    def test_rename_sheet(self, xlsx_path: str) -> None:
        excel_handler.excel_create_workbook(xlsx_path)
        excel_handler.excel_rename_sheet(xlsx_path, "Sheet", "数据表")
        result = excel_handler.excel_list_sheets(xlsx_path)
        assert "数据表" in result["sheets"]

    def test_copy_sheet(self, xlsx_path: str) -> None:
        excel_handler.excel_create_workbook(xlsx_path)
        excel_handler.excel_copy_sheet(xlsx_path, "Sheet", "副本")
        result = excel_handler.excel_list_sheets(xlsx_path)
        assert "副本" in result["sheets"]


class TestExcelDataReadWrite:
    def test_write_and_read_cell(self, xlsx_path: str) -> None:
        excel_handler.excel_create_workbook(xlsx_path)
        excel_handler.excel_write_cell(xlsx_path, "Sheet", "A1", "你好")
        result = excel_handler.excel_read_cell(xlsx_path, "Sheet", "A1")
        assert result["value"] == "你好"

    def test_write_and_read_range(self, xlsx_path: str) -> None:
        excel_handler.excel_create_workbook(xlsx_path)
        data = [["姓名", "年龄"], ["张三", 25], ["李四", 30]]
        excel_handler.excel_write_range(xlsx_path, "Sheet", "A1", data)
        result = excel_handler.excel_read_range(xlsx_path, "Sheet", "A1:B3")
        assert result["data"] == data

    def test_insert_delete_row(self, xlsx_path: str) -> None:
        excel_handler.excel_create_workbook(xlsx_path)
        excel_handler.excel_write_cell(xlsx_path, "Sheet", "A1", "原数据")
        excel_handler.excel_insert_row(xlsx_path, "Sheet", 1)
        result = excel_handler.excel_read_cell(xlsx_path, "Sheet", "A2")
        assert result["value"] == "原数据"


class TestExcelFormula:
    def test_apply_formula(self, xlsx_path: str) -> None:
        excel_handler.excel_create_workbook(xlsx_path)
        excel_handler.excel_write_cell(xlsx_path, "Sheet", "A1", 10)
        excel_handler.excel_write_cell(xlsx_path, "Sheet", "A2", 20)
        excel_handler.excel_apply_formula(xlsx_path, "Sheet", "A3", "=SUM(A1:A2)")
        result = excel_handler.excel_read_cell(xlsx_path, "Sheet", "A3")
        assert result["value"] == "=SUM(A1:A2)"


class TestExcelFreezePanes:
    def test_freeze(self, xlsx_path: str) -> None:
        excel_handler.excel_create_workbook(xlsx_path)
        result = excel_handler.excel_freeze_panes(xlsx_path, "Sheet", "B2")
        assert result["freeze_at"] == "B2"


class TestExcelAnalyzeData:
    def test_analyze(self, xlsx_path: str) -> None:
        excel_handler.excel_create_workbook(xlsx_path)
        data = [["姓名", "年龄"], ["张三", 25], ["李四", 30], ["王五", 35]]
        excel_handler.excel_write_range(xlsx_path, "Sheet", "A1", data)
        result = excel_handler.excel_analyze_data(xlsx_path, "Sheet", "A1:B4")
        assert result["total_rows"] == 3
        assert "年龄" in result["columns"]
        assert result["columns"]["年龄"]["type"] == "numeric"


class TestExcelFindDuplicates:
    def test_find_dups(self, xlsx_path: str) -> None:
        excel_handler.excel_create_workbook(xlsx_path)
        data = [["姓名"], ["张三"], ["李四"], ["张三"], ["王五"], ["张三"]]
        excel_handler.excel_write_range(xlsx_path, "Sheet", "A1", data)
        result = excel_handler.excel_find_duplicates(xlsx_path, "Sheet", ["姓名"])
        assert result["total_duplicates"] == 3
